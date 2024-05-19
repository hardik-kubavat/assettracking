#TODO - This class needs to be organized and also there should be a way to generate the log file.

from openpyxl import load_workbook
from project import logger
from project.extention import db
from project.models.ProductModel import Product
from project.models import ProductTypeModel,CategoryModel
from flask import url_for
import os

def import_products(file):
    work_book = load_workbook(file)
    sheet = work_book.active
    logger.debug("Column {} Rows ---{}".format(sheet.max_column,sheet.max_row-1))
    load_products(sheet)
    filepath = os.path.join('logs', file.filename)
    work_book.save(filepath)
    return filepath
    
def load_products(sheet):
    products=[]
    for row_obj in range(2,sheet.max_row+1):
        producttype = sheet.cell(row=row_obj,column=1).value
        category = sheet.cell(row=row_obj,column=2).value
        srno = sheet.cell(row=row_obj,column=3).value
        identification = sheet.cell(row=row_obj,column=4).value
        status = sheet.cell(row=row_obj,column=5).value
        owner = sheet.cell(row=row_obj,column=6).value
        remarks = sheet.cell(row=row_obj,column=7).value
        
        productTypeObj = ProductTypeModel.getByName(producttype)
        if productTypeObj is None:
            sheet.cell(row=row_obj,column=8).value = "Product Type Does not exists"
            logger.error("Product Type {} does not exists.".format(producttype))
            continue
        
        categoryObj = CategoryModel.getByNameAndProductType(category,productTypeObj.id)
        if categoryObj is None:
            sheet.cell(row=row_obj,column=8).value = "Category does not exists"
            logger.error("Category {} does not exists.".format(category))
            continue
        
        
        product_type_id = productTypeObj.id
        category_id = categoryObj.id
        try:
            product = Product(product_type_id=product_type_id, category_id=category_id, srno=srno, identification=identification, status=status, owner=owner, remarks=remarks)
            db.session.add(product)
            db.session.commit()
            sheet.cell(row=row_obj,column=8).value = "Success"
        except Exception as e:
            db.session.rollback()
            sheet.cell(row=row_obj,column=8).value = str(e)